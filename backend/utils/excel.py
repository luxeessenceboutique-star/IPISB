"""Génération de fichiers Excel (.xlsx) stylés, réutilisable par tous les
rapports comptables (paiements étudiants, instances fournisseurs, versements
bancaires…).

Objectif : reproduire fidèlement l'apparence des feuilles que les gérants
utilisent aujourd'hui (bandeau titre coloré, en-têtes en couleur, bordures,
format monétaire MAD) pour qu'ils retrouvent leurs repères.

Usage :
    from utils.excel import make_xlsx
    return make_xlsx(
        filename="Paiements_etudiants.xlsx",
        title="PAIEMENTS ÉTUDIANTS IPISB",
        theme="blue",
        columns=[
            {"key": "class_name", "label": "Filière / Promo"},
            {"key": "total_paye", "label": "Total payé", "type": "money"},
        ],
        rows=[{...}, ...],
    )
"""
from io import BytesIO

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Palettes calquées sur les feuilles Excel des gérants :
#   blue   → Paiements étudiants
#   green  → Versements bancaires AWB
#   yellow → Instances fournisseurs (échéancier)
_THEMES = {
    "blue":   {"banner": "2E86C1", "header": "AED6F1", "header_font": "1B4F72"},
    "green":  {"banner": "1E8449", "header": "ABEBC6", "header_font": "145A32"},
    "yellow": {"banner": "B7950B", "header": "FCF3A0", "header_font": "5B4B00"},
    "grey":   {"banner": "5D6D7E", "header": "D5DBDB", "header_font": "212F3D"},
}

_MONEY_FMT = '#,##0.00" MAD"'
_INT_FMT = "#,##0"


def make_xlsx(
    *,
    filename: str,
    title: str,
    columns: list[dict],
    rows: list[dict],
    theme: str = "blue",
    subtitle: str | None = None,
    sheet_name: str | None = None,
    legend: list[dict] | None = None,
) -> Response:
    """Construit un classeur .xlsx à une feuille et le renvoie en pièce jointe.

    columns : liste de dicts {key, label, type?, width?, align?, style?}
              type  ∈ {"text" (défaut), "money", "int", "date"}
              align ∈ {"left", "center", "right"} (surcharge l'alignement)
              style : callable(value, row) -> {"bg": hex, "fg": hex} | None
                      → colore la cellule de données selon sa valeur / sa ligne.
    rows    : liste de dicts indexés par `key`
    legend  : liste de dicts {label, bg, fg?} rendus en bas de feuille (pastille
              de couleur + libellé) pour expliquer le code couleur.
    """
    th = _THEMES.get(theme, _THEMES["blue"])
    ncol = max(1, len(columns))

    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or title)[:31]

    thin = Side(style="thin", color="BFC9CA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Ligne 1 : bandeau titre (fusionné sur toutes les colonnes) ──────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    banner = ws.cell(row=1, column=1, value=title)
    banner.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    banner.fill = PatternFill("solid", fgColor=th["banner"])
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    header_row = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font = Font(italic=True, size=10, color="616A6B")
        sub.alignment = Alignment(horizontal="right", vertical="center")
        header_row = 3

    # ── Ligne d'en-têtes ────────────────────────────────────────────────────
    header_fill = PatternFill("solid", fgColor=th["header"])
    header_font = Font(bold=True, color=th["header_font"])
    for j, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=j, value=col.get("label", col["key"]))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 24

    # ── Données ─────────────────────────────────────────────────────────────
    first_data = header_row + 1
    for i, row in enumerate(rows, start=first_data):
        for j, col in enumerate(columns, start=1):
            val = row.get(col["key"])
            typ = col.get("type", "text")
            align = col.get("align")
            cell = ws.cell(row=i, column=j)
            if typ == "money":
                cell.value = float(val or 0)
                cell.number_format = _MONEY_FMT
                cell.alignment = Alignment(horizontal=align or "right")
            elif typ == "int":
                cell.value = int(val or 0)
                cell.number_format = _INT_FMT
                cell.alignment = Alignment(horizontal=align or "center")
            elif typ == "date":
                cell.value = "" if val is None else str(val)
                cell.alignment = Alignment(horizontal=align or "center")
            else:
                cell.value = "" if val is None else str(val)
                cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
            cell.border = border
            # Coloration conditionnelle de la cellule (situation, statut, mois…).
            style_fn = col.get("style")
            if style_fn:
                st = style_fn(val, row)
                if st:
                    if st.get("bg"):
                        cell.fill = PatternFill("solid", fgColor=st["bg"])
                    if st.get("fg"):
                        cell.font = Font(color=st["fg"], bold=st.get("bold", False))

    # ── Finitions : figer les en-têtes, filtre auto, largeurs ───────────────
    ws.freeze_panes = ws.cell(row=first_data, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncol)}{header_row}"

    for j, col in enumerate(columns, start=1):
        width = col.get("width")
        if not width:
            longest = len(str(col.get("label", "")))
            for row in rows:
                v = row.get(col["key"])
                if v is not None:
                    longest = max(longest, len(str(v)))
            width = min(max(longest + 3, 12), 48)
        ws.column_dimensions[get_column_letter(j)].width = width

    # ── Légende (code couleur) ───────────────────────────────────────────────
    if legend:
        last_data = first_data + len(rows) - 1
        lg = max(last_data, header_row) + 2  # une ligne vide de séparation
        title_cell = ws.cell(row=lg, column=1, value="Légende")
        title_cell.font = Font(bold=True, color=th["header_font"])
        for k, item in enumerate(legend, start=1):
            r = lg + k
            swatch = ws.cell(row=r, column=1, value=item.get("sample", ""))
            swatch.fill = PatternFill("solid", fgColor=item["bg"])
            swatch.border = border
            swatch.alignment = Alignment(horizontal="center", vertical="center")
            if item.get("fg"):
                swatch.font = Font(color=item["fg"], bold=True)
            label = ws.cell(row=r, column=2, value=item.get("label", ""))
            label.alignment = Alignment(horizontal="left", vertical="center")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
