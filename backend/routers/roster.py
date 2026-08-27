"""Effectifs des stagiaires — reproduit le classeur "Canevas privé" (une
ligne par stagiaire : département/région/établissement/filière/année,
identité, CIN, Id massar…) comme une vraie table important/exportable,
INDÉPENDANTE des comptes de connexion (pas de auth.users par ligne — voir
sql/supabase_student_roster_migration.sql)."""
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile
from openpyxl import load_workbook
from supabase import Client

from deps import CurrentUser, get_current_user, get_db
from models import RosterCreate, RosterUpdate
from utils.audit import log_audit
from utils.excel import make_xlsx

router = APIRouter(prefix="/roster", tags=["roster"])

MAX_FILE_SIZE = 20 * 1024 * 1024  # 20 MB
XLSX_CONTENT_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",  # certains navigateurs/clients l'envoient ainsi
}

# (clé colonne DB, libellé attendu dans l'en-tête Excel — recherche insensible
# à la casse/aux espaces, dans cet ordre = ordre du fichier Canevas d'origine).
COLUMNS: list[tuple[str, str]] = [
    ("departement", "Département"),
    ("region", "Région"),
    ("province", "Province / préfecture"),
    ("milieu", "Urbain/Rural"),
    ("etablissement", "Etablissement"),
    ("mode_formation", "Mode de formation"),
    ("niveau_formation", "Niveau"),
    ("secteur", "Secteur"),
    ("filiere", "Filière"),
    ("annee_formation", "Année"),
    ("nom", "Nom"),
    ("prenom", "Prénom"),
    ("genre", "Genre"),
    ("besoins_specifiques", "Personne à besoins spécifiques"),
    ("type_handicap", "Type d'handicap"),
    ("cin", "CIN"),
    ("id_massar", "Id massar"),
    ("date_naissance", "Date de naissance"),
    ("nationalite", "Nationalité"),
    ("etranger_migrant_refugie", "Etranger/migrant/réfugié"),
    ("pays_origine", "Pays d'origine"),
    ("niveau_scolaire", "Niveau scolaire"),
    ("date_dernier_niveau", "Date du denier niveau scolaire"),
]
DATE_KEYS = {"date_naissance", "date_dernier_niveau"}


def _require_admin(user: CurrentUser) -> None:
    if not (user.is_admin() or user.is_rh() or user.is_assistant_rh()):
        raise HTTPException(403, "Admin ou RH uniquement")


def _norm(s) -> str:
    return "".join(ch for ch in str(s or "").strip().lower() if ch.isalnum())


def _parse_date(v) -> Optional[str]:
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.date().isoformat() if isinstance(v, datetime) else v.isoformat()
    s = str(v).strip()[:10]
    try:
        return date.fromisoformat(s).isoformat()
    except ValueError:
        return None


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    s = str(v).strip().lower()
    return s not in ("", "0", "non", "no", "false")


@router.get("")
async def list_roster(
    user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Client, Depends(get_db)],
    academic_year: Optional[str] = None,
    filiere: Optional[str] = None,
):
    _require_admin(user)
    query = db.from_("student_roster").select("*")
    if academic_year:
        query = query.eq("academic_year", academic_year)
    if filiere:
        query = query.eq("filiere", filiere)
    rows = query.order("filiere").order("annee_formation").order("nom").execute().data or []
    return rows


@router.get("/years")
async def list_years(
    user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Client, Depends(get_db)],
):
    _require_admin(user)
    rows = db.from_("student_roster").select("academic_year").execute().data or []
    return sorted({r["academic_year"] for r in rows if r.get("academic_year")}, reverse=True)


@router.post("")
async def create_roster_row(
    body: RosterCreate,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Client, Depends(get_db)],
):
    """Ajout manuel d'un stagiaire — même fiche qu'une ligne importée."""
    _require_admin(user)
    if not body.nom.strip() or not body.prenom.strip():
        raise HTTPException(400, "Nom et prénom sont obligatoires")
    row = body.model_dump()
    row["created_by"] = user.id
    res = db.from_("student_roster").insert(row).execute()
    log_audit(db, user.id, "roster.create", "student_roster", res.data[0]["id"])
    return res.data[0]


@router.patch("/{roster_id}")
async def update_roster_row(
    roster_id: str,
    body: RosterUpdate,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Client, Depends(get_db)],
):
    _require_admin(user)
    existing = db.from_("student_roster").select("id").eq("id", roster_id).execute().data
    if not existing:
        raise HTTPException(404, "Ligne introuvable")
    updates = body.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(400, "Aucun champ à mettre à jour")
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    res = db.from_("student_roster").update(updates).eq("id", roster_id).execute()
    log_audit(db, user.id, "roster.update", "student_roster", roster_id, updates)
    return res.data[0]


@router.post("/import")
async def import_roster(
    user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Client, Depends(get_db)],
    file: UploadFile,
    academic_year: str = "2025-2026",
    replace_existing: bool = False,
):
    """Importe un classeur au format Canevas (une feuille, un en-tête, une
    ligne par stagiaire). L'en-tête est repéré dynamiquement (recherche de
    la ligne contenant à la fois « Nom » et « Prénom ») pour tolérer un
    fichier légèrement différent du gabarit d'origine."""
    _require_admin(user)
    if file.content_type not in XLSX_CONTENT_TYPES and not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "Le fichier doit être un classeur Excel (.xlsx)")
    data = await file.read()
    if len(data) > MAX_FILE_SIZE:
        raise HTTPException(400, "Fichier trop volumineux (max 20 Mo)")
    if not data:
        raise HTTPException(400, "Fichier vide")

    try:
        wb = load_workbook(BytesIO(data), data_only=True)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        raise HTTPException(400, f"Impossible de lire le classeur : {e}")

    # Repère la ligne d'en-tête (contient "Nom" et "Prénom") sur les 10
    # premières lignes — le fichier d'origine a un bandeau titre + une ligne
    # vide avant l'en-tête réel.
    header_row_idx = None
    col_index: dict[str, int] = {}
    for r in range(1, min(11, ws.max_row) + 1):
        values = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        normed = {_norm(v): i for i, v in enumerate(values, start=1) if v}
        if _norm("Nom") in normed and _norm("Prénom") in normed:
            header_row_idx = r
            for key, label in COLUMNS:
                idx = normed.get(_norm(label))
                if idx:
                    col_index[key] = idx
            break

    if header_row_idx is None or "nom" not in col_index or "prenom" not in col_index:
        raise HTTPException(400, "En-tête introuvable — le fichier doit contenir au moins les colonnes « Nom » et « Prénom ».")

    inserted = 0
    skipped = 0
    rows_to_insert: list[dict] = []
    for r in range(header_row_idx + 1, ws.max_row + 1):
        nom = ws.cell(row=r, column=col_index["nom"]).value
        prenom = ws.cell(row=r, column=col_index.get("prenom", 0)).value if "prenom" in col_index else None
        if not (str(nom or "").strip()) or not (str(prenom or "").strip()):
            skipped += 1
            continue
        row: dict = {"academic_year": academic_year, "created_by": user.id}
        for key, _ in COLUMNS:
            idx = col_index.get(key)
            val = ws.cell(row=r, column=idx).value if idx else None
            if key in DATE_KEYS:
                row[key] = _parse_date(val)
            elif key == "besoins_specifiques":
                row[key] = _to_bool(val)
            elif key == "type_handicap":
                row[key] = None if val in (None, 0, "0") else str(val).strip()
            else:
                row[key] = None if val in (None, "") else str(val).strip()
        rows_to_insert.append(row)
        inserted += 1

    if not rows_to_insert:
        raise HTTPException(400, "Aucune ligne stagiaire exploitable trouvée dans le fichier.")

    if replace_existing:
        db.from_("student_roster").delete().eq("academic_year", academic_year).execute()

    db.from_("student_roster").insert(rows_to_insert).execute()
    log_audit(db, user.id, "roster.import", "student_roster", None,
              {"academic_year": academic_year, "inserted": inserted, "skipped": skipped, "replaced": replace_existing})
    return {"inserted": inserted, "skipped": skipped, "academic_year": academic_year}


@router.get("/export")
async def export_roster(
    user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Client, Depends(get_db)],
    academic_year: Optional[str] = None,
    filiere: Optional[str] = None,
):
    _require_admin(user)
    query = db.from_("student_roster").select("*")
    if academic_year:
        query = query.eq("academic_year", academic_year)
    if filiere:
        query = query.eq("filiere", filiere)
    rows = query.order("filiere").order("annee_formation").order("nom").execute().data or []

    columns = [{"key": key, "label": label} for key, label in COLUMNS]
    for row in rows:
        row["besoins_specifiques"] = "Oui" if row.get("besoins_specifiques") else "Non"

    year_label = academic_year or "toutes années"
    return make_xlsx(
        filename=f"Effectifs_stagiaires_{(academic_year or 'toutes').replace('/', '-')}.xlsx",
        title=f"EFFECTIFS DES STAGIAIRES — IPISB ({year_label})",
        theme="green",
        sheet_name=f"Stagiaires {academic_year or ''}".strip()[:31],
        columns=columns,
        rows=rows,
    )


@router.delete("/{roster_id}")
async def delete_roster_row(
    roster_id: str,
    user: Annotated[CurrentUser, Depends(get_current_user)],
    db: Annotated[Client, Depends(get_db)],
):
    _require_admin(user)
    existing = db.from_("student_roster").select("id").eq("id", roster_id).execute().data
    if not existing:
        raise HTTPException(404, "Ligne introuvable")
    db.from_("student_roster").delete().eq("id", roster_id).execute()
    log_audit(db, user.id, "roster.delete", "student_roster", roster_id)
    return {"ok": True}
